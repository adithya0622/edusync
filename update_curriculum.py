"""
Update Courses.xlsx with department-specific curriculum for all 8 sheets.
Run from: f:/edusync/Upgrade-main/
"""
import pandas as pd
from openpyxl import load_workbook

COURSES_FILE = 'backend/Courses.xlsx'

# Department-specific curriculum maps: {sheet_name: {assessment: curriculum}}
CURRICULUM = {
    # ── CSE 301: Computer Networks ──────────────────────────────────────────
    '19CSE301': {
        'Assignment 1': 'Introduction to the Network, The Internet, The Network Edge, Network Topology, Types of Networks',
        'Assignment 2': 'The Network Core, Circuit Switching and Packet Switching, Delay, Loss and Throughput in Packet Switched Networks',
        'Quiz 1': 'Protocol Layers and their Service Models, IP Stack, OSI Model',
        'Quiz 2': 'Introduction to Network Applications, The Web, HTTP, Persistent and Non-Persistent Connections',
        'Mid Lab Exam': 'Basic Networking Commands, Socket Programming in Java',
        'Mid Exam': 'Introduction to the Network, The Internet, Network Edge, Network Topology, Network Core, Circuit and Packet Switching, Protocol Layers, OSI Model',
        'Final Exam': 'All Units: Network Fundamentals, Packet Switching, OSI/IP Layers, HTTP, DNS, FTP, Email Protocols, SMTP, POP3, IMAP, Transport Layer, TCP, UDP, Congestion Control',
    },
    # ── CSE 302: Operating Systems ──────────────────────────────────────────
    '19CSE302': {
        'Assignment 1': 'Introduction to Operating Systems, Process Concepts, Process States, Process Control Block (PCB), Context Switching',
        'Assignment 2': 'CPU Scheduling Algorithms: FCFS, SJF, Round Robin, Priority Scheduling, Multilevel Queue Scheduling, Gantt Charts',
        'Quiz 1': 'Process Synchronization, Critical Section Problem, Mutex Locks, Semaphores, Monitors',
        'Quiz 2': 'Deadlocks: Necessary Conditions, Prevention, Avoidance, Detection and Recovery, Banker\'s Algorithm',
        'Quiz 3': 'Memory Management: Paging, Segmentation, Virtual Memory, Page Replacement Algorithms (FIFO, LRU, Optimal)',
        'Mid Lab Exam': 'CPU Scheduling Simulation using C/Java, POSIX Thread Programming, Semaphore Exercises',
        'Mid Exam': 'OS Introduction, Process Management, CPU Scheduling, Process Synchronization, Deadlock Handling',
        'Final Exam': 'Complete OS: Process Management, CPU Scheduling, Synchronization, Deadlocks, Memory Management, File Systems, I/O Management, Protection and Security',
    },
    # ── IT 301: Web Technologies ─────────────────────────────────────────────
    '19IT301': {
        'Assignment 1': 'Introduction to Web Technologies, HTML5 Semantic Elements, Forms, Tables, CSS3 Selectors, Box Model, Responsive Design Basics',
        'Assignment 2': 'JavaScript ES6+, DOM Manipulation, Event Handling, Promises, Async/Await, Fetch API, JSON',
        'Quiz 1': 'Server-side Programming with Node.js, Express.js, HTTP Request/Response Cycle, Middleware, REST API Design',
        'Quiz 2': 'Database Connectivity, MongoDB CRUD Operations, Mongoose ODM, AJAX, Session and Cookie Management',
        'Mid Lab Exam': 'Build a Responsive Web Application using HTML5, CSS3, and JavaScript with Backend Integration',
        'Mid Exam': 'HTML5, CSS3, JavaScript Fundamentals, Node.js, Express.js Basics, HTTP Protocol',
        'Final Exam': 'Full-Stack Web Development: Frontend (HTML/CSS/JS), Backend (Node.js/Express), Database (MongoDB), REST APIs, Web Security (XSS, CSRF), Deployment',
    },
    # ── IT 302: Database Management Systems ──────────────────────────────────
    '19IT302': {
        'Assignment 1': 'Introduction to DBMS, Database Architecture (1-Tier, 2-Tier, 3-Tier), Entity-Relationship Model, Entity Types, Attributes, Primary Keys',
        'Assignment 2': 'Relational Model, Relational Algebra (Select, Project, Join), SQL: DDL (CREATE, ALTER, DROP), DML (INSERT, UPDATE, DELETE)',
        'Quiz 1': 'Advanced SQL: Joins (Inner, Left, Right, Full Outer), Subqueries, Aggregation Functions, Views, Stored Procedures and Triggers',
        'Quiz 2': 'Database Normalization, Functional Dependencies, Closure, 1NF, 2NF, 3NF, BCNF, Lossless Decomposition',
        'Quiz 3': 'Transaction Management, ACID Properties, Serializability, Concurrency Control, Two-Phase Locking, Timestamp Ordering',
        'Mid Lab Exam': 'Database Design Project: ER Diagram to Relational Schema, SQL Query Optimization on real datasets',
        'Mid Exam': 'ER Modeling, Relational Model, SQL Commands, Normalization up to BCNF',
        'Final Exam': 'Complete DBMS: ER Design, SQL, Normalization, Transaction Management, Recovery Techniques, Indexing (B-Tree, Hash), Query Optimization',
    },
    # ── ADS 301: Python for Data Analytics ───────────────────────────────────
    '19ADS301': {
        'Assignment 1': 'Introduction to Python for Data Science, NumPy: Arrays, Matrix Operations, Broadcasting, Indexing and Slicing',
        'Assignment 2': 'Pandas: DataFrames, Series, Data Cleaning, Handling Missing Values, Data Merging and Reshaping',
        'Quiz 1': 'Descriptive Statistics: Mean, Median, Mode, Variance, Standard Deviation, Percentiles, Correlation, Covariance',
        'Quiz 2': 'Data Visualization: Matplotlib (Line, Bar, Histogram, Scatter), Seaborn (Heatmap, Pairplot, Box Plot), Plotly Basics',
        'Mid Lab Exam': 'Exploratory Data Analysis (EDA) on Real-World Dataset: cleaning, statistical summary, and visualization',
        'Mid Exam': 'Python Fundamentals, NumPy, Pandas DataFrames, Descriptive Statistics, Data Cleaning',
        'Final Exam': 'End-to-End Data Analytics Pipeline: Data Collection, Preprocessing, EDA, Statistical Analysis, Visualization, Interpretation and Reporting',
    },
    # ── ADS 302: Statistical Machine Learning ────────────────────────────────
    '19ADS302': {
        'Assignment 1': 'Introduction to Statistical Learning, Bias-Variance Tradeoff, Overfitting vs Underfitting, Training/Validation/Test Split',
        'Assignment 2': 'Linear Regression, Multiple Regression, Least Squares Method, Hypothesis Testing, R-Squared, Adjusted R-Squared',
        'Quiz 1': 'Classification: k-Nearest Neighbors (kNN), Naive Bayes Classifier, Logistic Regression, Decision Boundary, ROC Curve, AUC',
        'Quiz 2': 'Support Vector Machines (SVM), Kernel Functions (RBF, Polynomial), SVM for Classification and Regression (SVR)',
        'Quiz 3': 'Unsupervised Learning: K-Means Clustering, Hierarchical Clustering (Dendrogram), DBSCAN, Principal Component Analysis (PCA)',
        'Mid Lab Exam': 'Implement and evaluate Regression and Classification models using scikit-learn on real datasets',
        'Mid Exam': 'Statistical Learning Theory, Linear Regression, Classification Algorithms, Model Evaluation Metrics',
        'Final Exam': 'Full ML Pipeline: Preprocessing, Feature Engineering, Supervised/Unsupervised Learning, Model Selection, Hyperparameter Tuning, Cross-Validation, Evaluation',
    },
    # ── AIML 301: Introduction to Artificial Intelligence ────────────────────
    '19AIML301': {
        'Assignment 1': 'Introduction to Artificial Intelligence, Intelligent Agents, PEAS Framework, Types of Agents: Simple Reflex, Model-Based, Goal-Based, Utility-Based',
        'Assignment 2': 'Problem Solving by Search: State Space, BFS, DFS, Uniform Cost Search, Iterative Deepening DFS, Bidirectional Search',
        'Quiz 1': 'Informed Search: Heuristic Functions, Greedy Best-First Search, A* Algorithm, Admissibility and Consistency, Hill Climbing',
        'Quiz 2': 'Knowledge Representation: Propositional Logic, Wumpus World, First-Order Predicate Logic, Quantifiers, Inference Rules, Unification',
        'Mid Lab Exam': 'Implement BFS, DFS, A* Search Algorithms in Python; Solve 8-Puzzle and Maze problems',
        'Mid Exam': 'Intelligent Agents, Uninformed Search (BFS/DFS/UCS), Informed Search (A* /Greedy), Knowledge Representation',
        'Final Exam': 'Complete AI: Search Strategies, Logic, Knowledge Representation and Reasoning, Planning (STRIPS), Uncertainty (Bayesian Networks), Learning Basics',
    },
    # ── AIML 302: Deep Learning ───────────────────────────────────────────────
    '19AIML302': {
        'Assignment 1': 'Introduction to Neural Networks, Biological Inspiration, Perceptron, Multi-Layer Perceptron (MLP), Activation Functions (ReLU, Sigmoid, Tanh, Softmax), Forward Propagation',
        'Assignment 2': 'Backpropagation Algorithm, Computation Graph, Gradient Descent Variants (Batch, SGD, Mini-Batch), Optimizers: Adam, RMSprop, Momentum, Learning Rate Scheduling',
        'Quiz 1': 'Convolutional Neural Networks (CNN): Convolution Layers, Pooling, Feature Maps, Padding, Stride, Architectures: LeNet, AlexNet, VGG, ResNet',
        'Quiz 2': 'Recurrent Neural Networks (RNN): Vanishing Gradient Problem, Long Short-Term Memory (LSTM), Gated Recurrent Unit (GRU), Sequence-to-Sequence Models',
        'Quiz 3': 'Transfer Learning, Fine-Tuning Pre-trained Models, Generative Adversarial Networks (GANs), Variational Autoencoders (VAE)',
        'Mid Lab Exam': 'Build and Train a CNN using TensorFlow/Keras for Image Classification (MNIST/CIFAR-10)',
        'Mid Exam': 'Neural Network Fundamentals, Backpropagation, CNN Architectures, RNN/LSTM Applications',
        'Final Exam': 'Advanced Deep Learning: Attention Mechanism, Transformer Architecture, BERT, GPT Basics, Autoencoders, GANs, Model Deployment with TensorFlow Serving',
    },
}

def update_curriculum():
    wb = load_workbook(COURSES_FILE)
    
    for sheet_name, curriculum_map in CURRICULUM.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP: sheet {sheet_name} not found")
            continue
        
        ws = wb[sheet_name]
        
        # Find or create header row
        headers = [cell.value for cell in ws[1]]
        
        # Get column index for 'Assessments'
        if 'Assessments' not in headers:
            print(f"  SKIP: {sheet_name} has no Assessments column")
            continue
        assessment_col = headers.index('Assessments') + 1  # 1-based
        
        # Find or create 'Curriculum' column
        if 'Curriculum' in headers:
            curriculum_col = headers.index('Curriculum') + 1
        else:
            # Insert after last used column
            curriculum_col = len(headers) + 1
            # Find position after 'Converted Marks' or after 'Total Marks'
            for insert_after in ['Converted Marks', 'Total Marks']:
                if insert_after in headers:
                    curriculum_col = headers.index(insert_after) + 2  # after that col
                    # Insert a new column at curriculum_col
                    ws.insert_cols(curriculum_col)
                    ws.cell(row=1, column=curriculum_col).value = 'Curriculum'
                    # Update headers after insert
                    headers = [cell.value for cell in ws[1]]
                    break
            else:
                # Just append
                curriculum_col = len(headers) + 1
                ws.cell(row=1, column=curriculum_col).value = 'Curriculum'
        
        # Write curriculum for each assessment row
        for row in ws.iter_rows(min_row=2):
            assessment_val = row[assessment_col - 1].value
            if assessment_val and str(assessment_val).strip() in curriculum_map:
                row[curriculum_col - 1].value = curriculum_map[str(assessment_val).strip()]
        
        print(f"  Updated: {sheet_name}")
    
    wb.save(COURSES_FILE)
    print(f"\nSaved {COURSES_FILE}")

if __name__ == '__main__':
    print("Updating curriculum in Courses.xlsx...")
    update_curriculum()
    
    # Verify
    print("\n=== VERIFICATION ===")
    for sheet_name in CURRICULUM.keys():
        df = pd.read_excel(COURSES_FILE, sheet_name=sheet_name)
        if 'Curriculum' in df.columns:
            print(f"\n{sheet_name}:")
            for _, r in df.iterrows():
                cur = str(r['Curriculum'])[:80] if pd.notna(r.get('Curriculum')) else 'MISSING'
                print(f"  {r['Assessments']}: {cur}")
        else:
            print(f"\n{sheet_name}: NO CURRICULUM COLUMN!")
